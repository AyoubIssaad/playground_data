# from openpyxl import Workbook
import openpyxl
from xml.etree.ElementTree import Element, SubElement, Comment
from xml.etree import ElementTree
from xml.dom import minidom

def prettify(elem):
    rough_string = ElementTree.tostring(elem, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ")

items = Element("items")
comment = Comment('This will contain all your items')
items.append(comment)
wb = openpyxl.load_workbook(filename='source.xlsx')
ws = wb.active

# min_row = ws.min_row
# max_row = ws.max_row
# min_col = ws.min_column
# max_col = ws.max_column
# print(ws.max_column)
# print(ws.max_row)

for row_cells in ws.iter_rows():
    item = Element("item")
    name = SubElement(item, 'name')
    name.text = row_cells[1].value
    type = SubElement(item, 'type')
    type.text = 'TRAP'
    key = SubElement(item, 'key')
    key.text = row_cells[0].value
    delay = SubElement(item, 'delay')
    delay.text = '0'
    trends = SubElement(item, 'trends')
    trends.text = '0'
    value_type = SubElement(item, 'value_type')
    value_type.text = 'TEXT'
    allowed_hosts = SubElement(item, 'allowed_hosts')
    allowed_hosts.text = '192.168.30.0/24'
    triggers = SubElement(item, 'triggers')
    trigger1 = SubElement(triggers, 'trigger')
    expression1 = SubElement(trigger1, 'expression')
    expression1.text = '{last 1m)}&lt;&gt;0'
    name1 = SubElement(trigger1, 'name')
    name1.text = row_cells[2].value
    priority1 = SubElement(trigger1, 'priority')
    priority1.text = 'WARNING'
    dependencies1 = SubElement(trigger1, 'dependencies')
    dependency1 = SubElement(dependencies1, 'dependency')
    depName = SubElement(dependency1, 'name')
    depName.text = row_cells[3].value
    depExp = SubElement(dependency1, 'expression')
    depExp.text = row_cells[5].value


    trigger2 = SubElement(triggers, 'trigger')
    expression2 = SubElement(trigger2, 'expression')
    expression2.text = '{last(3m)}&lt;&gt;0'
    name2 = SubElement(trigger2, 'name')
    name2.text = row_cells[3].value
    priority2 = SubElement(trigger2, 'priority')
    priority1.text = 'HIGH'
    dependencies2 = SubElement(trigger2, 'dependencies')
    dependency2 = SubElement(dependencies2, 'dependency')
    depName2 = SubElement(dependency2, 'name')
    depName2.text = row_cells[4].value
    depExp2 = SubElement(dependency2, 'expression')
    depExp2.text = row_cells[6].value

    trigger3 = SubElement(triggers, 'trigger')
    expression3 = SubElement(trigger3, 'expression')
    expression3.text = '{last(5m)}&lt;&gt;0'
    name3 = SubElement(trigger3, 'name')
    name3.text = row_cells[4].value
    priority3 = SubElement(trigger3, 'priority')
    priority3.text = 'DISASTER'

    items.append(item)

text_file = open("items.xml", "w")
n = text_file.write(prettify(items))
text_file.close()
print(prettify(items))
# print(tostring(items))
    # print("*************************ROW*********************")
    # print(row_cells[0].value)
    # for cell in row_cells:
    #     print('%s: cell.value=%s' % (cell, cell.value) )
# for row in ws.iter_rows('A{}:A{}'.format(ws.min_row, ws.max_row)):
#     for cell in row:
#         print(cell.value)
# a1 = sheet['A1']
# print(a1.value)
# sheet_ranges = wb['range names']
# print(sheet_ranges['D18'].value)
